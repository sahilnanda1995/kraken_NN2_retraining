import lstm_neuralnet
import nn2
import requests as r
# from binance.client import Client
# from binance.websockets import BinanceSocketManager
import get_old_candle_data
from openpyxl import Workbook
from datetime import datetime
# client = Client('OEWj7QP2h6ROVHPWWSt58zB4NCg9tsjwhyzU5F0s7q469f6IAkL17vcg8AYKRGwO', 'CWj7tNqeNsTbzrjvIAdQPiRY3pEFQe4VWDlZK4LA2ndMxqpI4CIT1oezFR4VMkyD')
import krakenex
import time


book = Workbook()
sheet = book.active

count = 0
previous_price = 1000
takeAction = False
previous_action = 'no action'

def process_message(msg):
    #print("message type: {}".format(msg['e']))
    # print(msg)
    global count
    global previous_price
    global takeAction
    global previous_action
    global profit
    if msg[6] == True:
        #print(msg)
        closePrice = float(msg[4])
        lstm_neuralnet.appendLatestClose([float(msg[1]), float(msg[2]), float(msg[3]), float(msg[4])])
        count = count+1
        #print(count)
        if count == 10:
            count = 0
            takeAction = True
            #print('takeAction', takeAction)
            lstm_neuralnet.retrainingNN()
            profit = float(closePrice - previous_price)/previous_price * 100
        
        #printMethod.printCandle('hello there')
        #printMethod.pushIntoX(float(msg['k']['c']))
        #print('hello there last candle :)')
        predictedPrice = lstm_neuralnet.predict_value(float(msg[4]))
        percDiff = float(predictedPrice - closePrice) / closePrice * 100
        if takeAction == True:
            print('here')
            print('volume', float(msg[5]))
            not_skipping_prob = nn2.predict_probab_not_skipping([float(msg[4])], [predictedPrice], [float(msg[5])])
            print('not_skipping_prob', not_skipping_prob)
            if percDiff > 0:
                if previous_action == 'long':
                    print('profit', profit)
                    sheet.cell(row = sheet.max_row, column = 7, value = previous_price)
                    sheet.cell(row = sheet.max_row, column = 8, value = profit)
                    sheet.cell(row = sheet.max_row, column = 9, value = profit-0.2)
                elif previous_action == 'short':
                    print('profit', -profit)
                    sheet.cell(row = sheet.max_row, column = 7, value = previous_price)
                    sheet.cell(row = sheet.max_row, column = 8, value = profit)
                    sheet.cell(row = sheet.max_row, column = 9, value = profit-0.2)
                print('Price', closePrice, 'Predicted Price', predictedPrice ,'percDiff', percDiff, 'Action', 'long')
                sheet.append((datetime.utcfromtimestamp(msg[0]).strftime('%Y-%m-%d %H:%M:%S'), closePrice, predictedPrice,  percDiff, not_skipping_prob, 'long'))
                book.save('appendingKraken.xlsx')
                previous_action = 'long'
            elif percDiff <= 0:
                if previous_action == 'long':
                    print('profit', profit)
                    sheet.cell(row = sheet.max_row, column = 7, value = previous_price)
                    sheet.cell(row = sheet.max_row, column = 8, value = profit)
                    sheet.cell(row = sheet.max_row, column = 9, value = profit-0.2)
                elif previous_action == 'short':
                    print('profit', -profit)
                    sheet.cell(row = sheet.max_row, column = 7, value = previous_price)
                    sheet.cell(row = sheet.max_row, column = 8, value = profit)
                    sheet.cell(row = sheet.max_row, column = 9, value = profit-0.2)
                print('Price', closePrice, 'Predicted Price', predictedPrice ,'percDiff', percDiff, 'Action', 'short')
                sheet.append((datetime.utcfromtimestamp(msg[0]).strftime('%Y-%m-%d %H:%M:%S'), closePrice, predictedPrice,  percDiff, not_skipping_prob, 'short'))
                book.save('appendingKraken.xlsx')
                previous_action = 'short'
            previous_price = closePrice
        takeAction = False

# print(client.ping())
#print(client.get_exchange_info())
# print(client.get_server_time())

#klines = client.get_historical_klines("BTCUSDT", '5m', "1 Jan, 2019")
#y = list(map(lambda x: [float(x[4])], klines))
#print(y[0:10])

# arrX = get_old_candle_data.last_testArr()
# print(len(arrX))
# lstm_neuralnet.predict_value(arrX)
# bm = BinanceSocketManager(client)
# bm.start_kline_socket('BTCUSDT', process_message, interval='5m')
# bm.start()

while True:
    #print(time.time())
    t = int(time.time())
    #print(t)
    if t%300 == 0:
        #print(time.gmtime(t))
        xxbt = r.get('https://api.kraken.com/0/public/OHLC?pair=XBTUSD&interval=5').json()
        #print(xxbt)
        for keys, vals in xxbt.items():
            #print(vals)
            #print(type(vals))
            if keys == 'result':
                #print(vals)
                for key, val in vals.items():
                    if type(val) == list:
                        #print(val[len(val)-2])
                        y = list(map(lambda x: [float(x[0]),float(x[1]),float(x[2]),float(x[3]),float(x[4]),float(x[6]), True], val))
                        #print('liveData', y[len(y)-2])
                        process_message(y[len(y)-2])

    #print("tick")
    time.sleep(1)
